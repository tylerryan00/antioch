from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import math
import random

help_description = """
Random Groups

This script has the goal of randomizing given groups
"""


def randomizer(names, current_name):
    return names[current_name]


def workbook_setup(wb, day, names, van_size):
    worksheet_name = 'Day ' + str(day)
    ws = wb.create_sheet(worksheet_name)
    thin = Side(border_style="thin", color="000000")
    y = 0
    van_count = 0
    random.shuffle(names)

    for x in range(len(names)):
        if x % van_size == 0:
            y += 2
            van_count += 1
            ws.merge_cells(start_row=1, start_column=y,
                           end_row=1, end_column=y+1)
            cell = ws.cell(row=1, column=y, value="Van " + str(van_count))
            adj_cell = ws.cell(row=1, column=y+1)

            # format cells
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            adj_cell.border = Border(
                top=thin, left=thin, right=thin, bottom=thin)
            ws.column_dimensions[chr(y+63)].width = 4
            ws.column_dimensions[chr(y+64)].width = 15
            ws.column_dimensions[chr(y+65)].width = 15

        elif x % math.ceil(van_size/2) == 0:
            y += 1

        ws.cell(row=int((x % (van_size/2))+2), column=y, value=randomizer(names, x)
                ).border = Border(top=thin, left=thin, right=thin, bottom=thin)


def main():
    # read in the list of names
    with open('names.txt') as f:
        names = f.readlines()

    # test cases
    days_of_trip = 5
    max_van_size = 10

    # find the optimal
    num_kids = len(names)
    vans_needed = math.ceil(num_kids / max_van_size)
    optimal_van_size = math.ceil(num_kids / vans_needed)

    # Creates the workbook and first sheet 
    # Makes first sheet a list of names
    workbook = Workbook()
    ws = workbook.active
    ws.title = 'People'
    for x in range(len(names)):
        ws.cell(row=x+1, column=1, value=names[x])

    # Creates the next sheets based on amount
    # of days of trip, plus adds the data
    for day in range(1, days_of_trip+1):
        workbook_setup(workbook, day, names, optimal_van_size)

    workbook.save('test.xlsx')


if __name__ == "__main__":
    main()
